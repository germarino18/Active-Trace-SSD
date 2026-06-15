"""File parsing helpers for padrón import (C-09 TASK GROUP 5.3).

Supports .xlsx (openpyxl) and .csv (csv module) files. Normalizes column
names to lowercase with stripped whitespace. Validates required columns:
nombre, apellidos, email. Optional: comision, regional.
"""

import csv
import io
import os

from app.core.exceptions import ValidationException

REQUIRED_COLUMNS = {"nombre", "apellidos", "email"}
OPTIONAL_COLUMNS = {"comision", "regional"}


def _normalize_header(h: str) -> str:
    return h.strip().lower()


def _validate_columns(column_names: list[str]) -> None:
    normalized = {_normalize_header(c) for c in column_names}
    missing = REQUIRED_COLUMNS - normalized
    if missing:
        raise ValidationException(
            message=f"Columnas requeridas faltantes: {', '.join(sorted(missing))}",
            details={"missing_columns": sorted(missing), "found_columns": column_names},
        )


def _parse_csv(file_content: bytes) -> tuple[list[dict], list[str]]:
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValidationException(message="El archivo CSV está vacío o no tiene encabezados")

    raw_column_names = list(reader.fieldnames)
    rows = []
    for row in reader:
        normalized = {_normalize_header(k): v for k, v in row.items()}
        rows.append(normalized)

    normalized_columns = [_normalize_header(c) for c in raw_column_names]
    return rows, normalized_columns


def _parse_xlsx(file_content: bytes) -> tuple[list[dict], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValidationException(message="El archivo XLSX no tiene hojas de cálculo")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise ValidationException(message="El archivo XLSX está vacío")

    if raw_headers is None:
        raise ValidationException(message="El archivo XLSX no tiene encabezados")

    raw_column_names = [str(h) if h is not None else "" for h in raw_headers]
    normalized_columns = [_normalize_header(c) for c in raw_column_names]
    rows = []
    for row_values in rows_iter:
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(normalized_columns):
                col = normalized_columns[i]
                if col:
                    row_dict[col] = str(val) if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return rows, normalized_columns


def parse_padron_file(file_content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Parse xlsx or csv file, return (rows_as_dicts, column_names).

    Column names are normalized to lowercase with stripped whitespace.
    Raises ValidationException if required columns are missing or format
    is unsupported.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        rows, column_names = _parse_csv(file_content)
    elif ext in (".xlsx", ".xls"):
        rows, column_names = _parse_xlsx(file_content)
    else:
        raise ValidationException(
            message=f"Formato de archivo no soportado: {ext}. Use .csv o .xlsx",
            details={"extension": ext, "filename": filename},
        )

    _validate_columns(column_names)
    return rows, column_names
