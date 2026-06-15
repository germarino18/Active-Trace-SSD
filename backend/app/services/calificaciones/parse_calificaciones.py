"""File parsing helpers for calificaciones import (C-10).

Supports .xlsx (openpyxl) and .csv (csv module) files. Detects grade
columns with (Real) suffix (numeric) and non-(Real) columns (textual).

RN-01: Columnas que terminan en "(Real)" son nota numérica.
RN-02: Columnas textuales contienen valores como "Satisfactorio".
"""

import csv
import io
import os
import re


def _normalize_header(h: str) -> str:
    return h.strip()


def _parse_csv(file_content: bytes) -> tuple[list[dict], list[str]]:
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("El archivo CSV está vacío o no tiene encabezados")

    raw_column_names = list(reader.fieldnames)
    rows = []
    for row in reader:
        rows.append(row)
    return rows, raw_column_names


def _parse_xlsx(file_content: bytes) -> tuple[list[dict], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("El archivo XLSX no tiene hojas de cálculo")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise ValueError("El archivo XLSX está vacío")

    if raw_headers is None:
        raise ValueError("El archivo XLSX no tiene encabezados")

    raw_column_names = [str(h) if h is not None else "" for h in raw_headers]
    rows = []
    for row_values in rows_iter:
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(raw_column_names):
                col = raw_column_names[i]
                if col:
                    row_dict[col] = str(val) if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return rows, raw_column_names


def _parse_file(file_content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        return _parse_csv(file_content)
    elif ext in (".xlsx", ".xls"):
        return _parse_xlsx(file_content)
    else:
        raise ValueError(
            f"Formato de archivo no soportado: {ext}. Use .csv o .xlsx"
        )


_REAL_SUFFIX_RE = re.compile(r"^(.*)\s*\(Real\)$", re.IGNORECASE)

_STUDENT_COLUMNS = {"apellidos", "nombre"}


def _is_student_column(name: str) -> bool:
    return name.strip().lower() in _STUDENT_COLUMNS


def _try_parse_float(val: str) -> float | None:
    if val is None or val.strip() == "":
        return None
    try:
        return float(val.replace(",", "."))
    except (ValueError, TypeError):
        return None


def _get_row_value(row: dict, col: str) -> str | None:
    val = row.get(col)
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    return str(val).strip()


def parse_calificaciones_file(file_content: bytes, filename: str) -> dict:
    """Parse LMS grade file and detect columns.

    Returns: {
        "actividades": [{"nombre", "tipo", "columna"}],
        "filas": [{"alumno_nombre", "alumno_apellidos", ...activity_columns...}],
        "total_filas": int,
    }
    """
    rows, column_names = _parse_file(file_content, filename)

    actividades = []
    activity_keys: set[tuple[str, str]] = set()

    for col in column_names:
        normalized = _normalize_header(col)
        if _is_student_column(normalized):
            continue

        match = _REAL_SUFFIX_RE.match(normalized)
        if match:
            activity_name = match.group(1).strip()
            key = (activity_name, "numerica")
            if key not in activity_keys:
                activity_keys.add(key)
                actividades.append({
                    "nombre": activity_name,
                    "tipo": "numerica",
                    "columna": normalized,
                })
        else:
            name = normalized
            key = (name, "textual")
            if key not in activity_keys:
                activity_keys.add(key)
                actividades.append({
                    "nombre": name,
                    "tipo": "textual",
                    "columna": normalized,
                })

    parsed_rows = []
    for row in rows:
        parsed = {}
        parsed["alumno_nombre"] = row.get("Nombre", row.get("nombre", ""))
        parsed["alumno_apellidos"] = row.get("Apellidos", row.get("apellidos", ""))

        for act in actividades:
            act_name = act["nombre"]
            col = act["columna"]

            if act["tipo"] == "numerica":
                real_col = f"{act_name} (Real)"
                numeric_val = _try_parse_float(_get_row_value(row, real_col) or "")
                parsed[f"{act_name}_num"] = numeric_val
            else:
                textual_val = _get_row_value(row, col)
                parsed[f"{act_name}_txt"] = textual_val

        parsed_rows.append(parsed)

    return {
        "actividades": actividades,
        "filas": parsed_rows,
        "total_filas": len(parsed_rows),
    }


def parse_finalizacion_file(file_content: bytes, filename: str) -> dict:
    """Parse completion report file.

    Similar to parse_calificaciones_file but detects columns with
    completion status (Aprobado/No aprobado/Entregado).
    """
    rows, column_names = _parse_file(file_content, filename)

    actividades = []
    for col in column_names:
        normalized = _normalize_header(col)
        if _is_student_column(normalized):
            continue
        actividades.append({
            "nombre": normalized,
            "tipo": "finalizacion",
            "columna": normalized,
        })

    parsed_rows = []
    for row in rows:
        parsed = {}
        parsed["alumno_nombre"] = row.get("Nombre", row.get("nombre", ""))
        parsed["alumno_apellidos"] = row.get("Apellidos", row.get("apellidos", ""))

        for act in actividades:
            val = _get_row_value(row, act["columna"])
            parsed[act["nombre"]] = val

        parsed_rows.append(parsed)

    return {
        "actividades": actividades,
        "filas": parsed_rows,
        "total_filas": len(parsed_rows),
    }


def detectar_tps_sin_calificar(rows: list[dict], actividades: list[dict]) -> list[dict]:
    """Detect TPs that are 'Entregado' but have no numeric grade.

    Returns list of {actividad, alumno_nombre, alumno_apellidos}.
    """
    result = []
    for row in rows:
        for act in actividades:
            act_name = act["nombre"]
            txt_key = f"{act_name}_txt"
            num_key = f"{act_name}_num"

            textual_val = row.get(txt_key)
            numeric_val = row.get(num_key)

            if textual_val == "Entregado" and numeric_val is None:
                result.append({
                    "actividad": act_name,
                    "alumno_nombre": row.get("alumno_nombre", ""),
                    "alumno_apellidos": row.get("alumno_apellidos", ""),
                })
    return result
